# -*- coding: utf-8 -*-

##########差值##########
from osgeo import ogr
import os
from osgeo import gdal
from openpyxl import Workbook
import pandas as pd


#1——原来的
#2——新加的

#读取文件
#NDVIpath=r'F:\Data\outndvi'
shuiweipath=r'F:\Conherence of InSAR with MachineLearning\Data\waterlevelfromgaugestations\shuiwei12'
copath=r'F:\Conherence of InSAR with MachineLearning\Data\Resample\covv'
temppath=r'F:\Conherence of InSAR with MachineLearning\Data\Resample\t2m'
uwindpath=r'F:\Conherence of InSAR with MachineLearning\Data\Resample\u_wind'
vwindpath=r'F:\Conherence of InSAR with MachineLearning\Data\Resample\v_wind'
rainfallpath=r'F:\Conherence of InSAR with MachineLearning\Data\Resample\total precipitation'
inputSHP = r'F:\Conherence of InSAR with MachineLearning\Data\study area\baohuqu\p_wca2a.shp'# 点的shp
#saveExcelPathAndName = r'F:\Data\zhandianexcel' # Excel保存的位置
zhandianpath = r'F:\Conherence of InSAR with MachineLearning\Data\study area\baohuqu\p_wca2a.xls'# Excel保存的位置

# 获取列表的第元素
#def take(elem):
    #return elem[2:10]

#获取文件并排序
#def getfile(path):
    #input_folder_list=os.listdir(path) # 读取文件夹里所有文件
    #tif_files=[] # 创建一个只装tif格式的列表
    #for filename in input_folder_list:  #遍历
     #   if os.path.splitext(filename)[1] == '.tif':  # 不管文件名里面有多少个tif，都只认最后一个tif
     #       tif_files.append(filename)  # 将文件夹里的tif文件加入只有tif的列表
    #tif_files.sort(key=take)
    #print(tif_files)
    #return tif_files

#获取一个目录下的所有文件
def get_filelist(path):
    Filelist = []
    for home, dirs, files in os.walk(path):
        for filename in files:
            # 文件名列表，包含完整路径
            Filelist.append(os.path.join(home, filename))
            # # 文件名列表，只包含文件名
            # Filelist.append( filename)
    return Filelist

# 获取列表的第几个元素
def take(elem):
    return elem[-12:-4]

#获取文件并排序
def getfile(file):
    tif_files=[] # 创建一个只装tif格式的列表
    #print(len(file))
    #for filename in file:
        #print(filename[-4:])
    for filename in file:  #遍历
        if filename[-4:-1] == '.ti':  # 不管文件名里面有多少个tif，都只认最后一个tif
            tif_files.append(filename)  # 将文件夹里的tif文件加入只有tif的列表            
    tif_files.sort(key=take)
    #for file in tif_files:
        #print(tif_files)
    return tif_files

#获取矢量点位的经纬度
def getpoint(path):
    # 设置driver
    driver = ogr.GetDriverByName('ESRI Shapefile')
    # 打开矢量
    ds = driver.Open(path, 0)
    # 获取图层
    layer = ds.GetLayer()
    # 获取要素及要素地理位置
    xValues = []
    yValues = []
    feature = layer.GetNextFeature()
    while feature:
        #shprow = shprows.next()        
        geometry = feature.GetGeometryRef()
        x = geometry.GetX()
        y = geometry.GetY()
        #zhandian.append(shprow.name)
        xValues.append(x)
        yValues.append(y)
        feature = layer.GetNextFeature()
    return xValues,yValues


#获取点在栅格中的位置和其值
def getposition(tif_files,xValues,yValues):
     ds = gdal.Open(tif_files, gdal.GA_ReadOnly)
     #print(ds)
     #print(list[i])
     # 获取行列、波段
     #rows = ds.RasterYSize
     #cols = ds.RasterXSize
     #bands = ds.RasterCount
     # 获取放射变换信息
     transform = ds.GetGeoTransform()
     #print(transform)
     xOrigin = transform[0]
     yOrigin = transform[3]
     pixelWidth = transform[1]
     pixelHeight = transform[5]
     #
     value =[]
     for j in range(len(xValues)):       # 遍历所有点
         x = xValues[j]
         y = yValues[j]
         #获取点位所在栅格的位置
         xOffset = int((x - xOrigin) / pixelWidth)
         yOffset = int((y - yOrigin) / pixelHeight)
         band = ds.GetRasterBand(1)
         data = band.ReadAsArray(xOffset, yOffset, 1, 1) # 读出从(xoffset,yoffset)开始，大小为(xsize,ysize)的矩阵
         value.append(str(data[0, 0]))
     return value


##计算两天差值##
def dvalue(tif_files1,tif_files2,xValues,yValues):
    dvalue1=[]
    dvalue2=[]
    dvalue=[]
    #print(tif_files1)
    #print(tif_files2)
    dvalue1=getposition(tif_files1,xValues,yValues)
    dvalue2=getposition(tif_files2,xValues,yValues)
    #print(len(dvalue1))
    #print(len(dvalue2))
    for k in range(len(dvalue1)):
        dvalue.append(float(dvalue2[k])-float(dvalue1[k]))
    return dvalue
    

if __name__ == "__main__":
    ##获取站点名
    df = pd.read_excel(zhandianpath)
    zdname=df["Name"]
    length=len(zdname)
    print("站点总数：",length)
    #获取文件
    #c_file_list = getfile(copath)
    #s_file_list = getfile(shuiweipath)
    #t_file_list = getfile(temppath)
    #wu_file_list = getfile(uwindpath)
    #wv_file_list = getfile(vwindpath)
    #获取文件
    c_file = get_filelist(copath)
    c_file_list = getfile(c_file)
    print(len(c_file_list))
    s_file = get_filelist(shuiweipath)
    s_file_list = getfile(s_file)
    print(len(s_file_list))
    t_file = get_filelist(temppath)
    t_file_list = getfile(t_file)
    print(len(t_file_list))
    u_file = get_filelist(uwindpath)
    wu_file_list = getfile(u_file)
    print(len(wu_file_list))
    v_file = get_filelist(vwindpath)
    wv_file_list = getfile(v_file)
    print(len(wv_file_list))
    r_file = get_filelist(rainfallpath)
    r_file_list = getfile(r_file)
    print(len(r_file_list))
    #n_file_list = getfile(NDVIpath)
    #list=[copath,shuiweipath,temppath,uwindpath,vwindpath]
    #,temppath,uwindpath,vwindpath
    xValues,yValues=getpoint(inputSHP)
    ##新建工作簿
    wb = Workbook()
    #新建工作表并修改名称#2
    ws = wb.active#2
    for i in range(len(c_file_list)):
        #获取文件
        print("第",i,"次提取")
        tif_files=[]
        tif_files.append(c_file_list[i])
        tif_files.append(s_file_list[i])
        tif_files.append(s_file_list[i+1])
        tif_files.append(t_file_list[i])
        tif_files.append(t_file_list[i+1])
        tif_files.append(wu_file_list[i])
        tif_files.append(wu_file_list[i+1])
        tif_files.append(wv_file_list[i])
        tif_files.append(wv_file_list[i+1])
        tif_files.append(r_file_list[i])
        tif_files.append(r_file_list[i+1])
        #filename=c_file_list[i]
        #filename2=c_file_list[i+1]
        #tif_files.append(n_file_list[i])
        #tif_files.append(n_file_list[i+1])
        #新建工作表并修改名称1
        #ws = wb.active#1
        #filename=tif_files[0]#1
        #ws.title =filename[2:10]#1
        # 写入标题
        ws.cell(row=1, column=1, value ='zhandian')
        ws.cell(row=1, column=2, value ='X')
        ws.cell(row=1, column=3, value ='Y')
        ws.cell(row=1, column=4, value ='coherence')
        ws.cell(row=1, column=5, value ='waterlevel')
        ws.cell(row=1, column=6, value ='temperature')
        ws.cell(row=1, column=7, value ='u_windpower')
        ws.cell(row=1, column=8, value ='v_windpower')
        ws.cell(row=1, column=9, value ='rainfall')
        ws.cell(row=1, column=10, value ='year')#2
        ws.cell(row=1, column=11, value ='month')#2
        ws.cell(row=1, column=12, value ='day')#2
        filename = tif_files[0]
        print("时间为：",filename)
        #ws.cell(row=1, column=8, value ='NDVI')
        #n=0  #设置一个变量用来"记"文件的第几个
        #计算点位
        for j in range(6):
            if(j!=0):
                value1=dvalue(tif_files[2*j-1],tif_files[2*j],xValues,yValues)
            else:
                value1=[]
                value1=getposition(tif_files[j],xValues,yValues)

            for k in range(len(xValues)):
                ws.cell(row=k+2+i*length, column=1, value = zdname[k])
                ws.cell(row=k+2+i*length, column=2, value = xValues[k])
                ws.cell(row=k+2+i*length, column=3, value = yValues[k])
                ws.cell(row=k+2+i*length, column=4+j, value = value1[k])
                ws.cell(row=k+2+i*length, column=10, value =filename[-12:-8])#2
                ws.cell(row=k+2+i*length, column=11, value =filename[-8:-6])#2
                ws.cell(row=k+2+i*length, column=12, value =filename[-6:-4])#2
                #print(zdname[k])
            #n=n+1
            if(j+1==len(tif_files)):
                    break
    wb.save(r'F:\Conherence of InSAR with MachineLearning\Data\自变量和因变量\数据提取\按保护区位置\\'+ 'dvalue_WCA2A_' + '.xlsx')

















